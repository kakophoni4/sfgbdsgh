from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Без букв A–X в шапке — понятные названия для заказчика
HEADERS = [
    "Название",
    "ИНН",
    "Дата регистрации",
    "Цена",
    "Цена покупки (вручную)",
    "Адрес и директор",
    "Налоговый режим",
    "Регистрация до 2024",
    "Достоверность ЕГРЮЛ",
    "Первичка / 1С (вручную)",
    "Займы и кредиторка",
    "Долги / исполнительные листы",
    "Не на ликвидации",
    "Не на исключении",
    "Банкротство",
    "Судебные дела",
    "Расчётные счета (вручную)",
    "Обороты (>500 тыс.)",
    "ЗСК (заявление продавца)",
    "Приостановки (вручную)",
    "Отчётность сдана",
    "Лизинг / залоги",
    "Первое появление в чате",
    "Дата проверки",
    "Итог: брать или нет",
    "Ссылка на объявление",
    "Продавец",
    "Балл",
    "ОГРН",
    "ОКВЭД",
    "Уверенность",
    "Статус в ЕГРЮЛ",
    "Дубликат",
    "Карточка Companium",
    "Сотрудники",
    "МСП",
    "Санкции",
    "Уставный капитал",
    "Налоги (уплачено)",
    "Учредитель",
    "Сырой текст",
]

ZSK_FILL = {
    "green": PatternFill("solid", fgColor="C6EFCE"),
    "yellow": PatternFill("solid", fgColor="FFEB9C"),
    "red": PatternFill("solid", fgColor="FFC7CE"),
}

VERDICT_FILL = {
    "ДА": PatternFill("solid", fgColor="C6EFCE"),
    "НЕТ": PatternFill("solid", fgColor="FFC7CE"),
    "СОМНИТЕЛЬНО": PatternFill("solid", fgColor="FFEB9C"),
}


def _zsk_cell(claim: str) -> str:
    return {
        "green": "зелёный (заявление продавца)",
        "yellow": "жёлтый (заявление продавца)",
        "red": "красный (заявление продавца)",
        "unknown": "не указан в посте",
        "": "не указан в посте",
    }.get(claim or "unknown", "не указан в посте")


def _zsk_cell_payload(p: dict[str, Any]) -> str:
    """Бот главнее заявления продавца."""
    bot = ((p.get("enrich") or {}).get("zsk_bot") or {})
    level = (p.get("zsk_verified") or bot.get("level") or "").strip()
    claim = p.get("zsk_claim") or "unknown"
    if level in {"green", "yellow", "red"} and not bot.get("error"):
        color = {"green": "зелёный", "yellow": "жёлтый", "red": "красный"}[level]
        label = str(bot.get("label") or {"green": "Низкий", "yellow": "Средний", "red": "Высокий"}[level])
        extra = ""
        if claim in {"green", "yellow", "red"} and claim != level:
            extra = f"; в посте: {_zsk_cell(claim).split('(')[0].strip()}"
        elif claim == level:
            extra = "; пост совпадает"
        return f"{color} (бот): {label}{extra}"
    if bot.get("error"):
        return f"{_zsk_cell(claim)}; бот: {bot.get('error')}"
    return _zsk_cell(claim)


def _r_cell(checklist: dict[str, Any], scoring: dict[str, Any], p: dict[str, Any]) -> str:
    r = str(checklist.get("R_turnover") or "").strip()
    buh_ok = bool(
        ((p.get("enrich") or {}).get("buh") or {}).get("years")
        and not ((p.get("enrich") or {}).get("buh") or {}).get("error")
    )
    if r:
        if buh_ok or r.startswith("есть") or r.startswith("мало") or "выручка" in r or "БФО" in r:
            return r if r.startswith("БФО") else f"БФО: {r}"
        return r
    # без БФО — только пометка про текст поста, не притворяемся реестром
    if p.get("revenues"):
        ht = scoring.get("has_turnover_flag") or "есть цифры"
        return f"только в посте (не БФО): {ht}"
    if p.get("zero_turnover_claim"):
        return "продавец: без оборотов / нулёвка"
    return "нет данных (БФО пусто / не найдено)"


def _price_cell(price: int | None) -> str | int:
    return price if price is not None else ""


def _money(v: Any) -> str:
    if v is None or v == "":
        return ""
    try:
        return f"{int(v):,}".replace(",", " ") + " ₽"
    except (TypeError, ValueError):
        return str(v)


def _human_flag(key: str, val: Any) -> str:
    """Внутренние ДА/НЕТ → понятные фразы в Excel."""
    s = str(val or "").strip()
    if not s:
        return ""
    low = s.lower()
    if key == "I_reliable":
        if low == "да":
            return "достоверно (записи о недостоверности нет)"
        if low == "нет":
            return "есть недостоверность сведений"
        if "провер" in low:
            return "нужно проверить"
    if key in {"M_not_liquidating", "N_not_excluding"}:
        if low == "да":
            return "да"
        if low == "нет":
            return "нет — риск"
    if key == "U_reports_filed":
        if low == "да":
            return "сдана"
        if low == "нет":
            return "не найдена"
    if s == "ПРОВЕРИТЬ":
        return "нужно проверить"
    return s


def _v_cell(checklist: dict[str, Any]) -> str:
    """Лизинг/Федресурс: текст + ссылка, без голого «проверить»."""
    v = _human_flag("V_leases", checklist.get("V_leases"))
    raw = str(checklist.get("V_leases") or "").lower()
    if "есть записи" in raw:
        v = "есть записи на Федресурсе"
    elif raw in {"нет лизинга/залогов", "нет"}:
        v = "нет лизинга/залогов"
    elif "есть лизинг" in raw:
        v = "есть лизинг/залоги"
    link = (checklist.get("V_link") or "").strip()
    note = (checklist.get("V_note") or "").strip()
    parts = [p for p in (v, note, link) if p]
    return " | ".join(parts)


def row_from_payload(p: dict[str, Any]) -> list[Any]:
    scoring = p.get("scoring") or {}
    enrich = p.get("enrich") or {}
    checklist = enrich.get("checklist") or {}
    egrul = enrich.get("egrul") or {}
    companium = enrich.get("companium") or {}

    j_hint = ""
    if p.get("primary_1c_claim"):
        j_hint = "продавец обещает первичку/1С"
    q_hint = ""
    if p.get("has_account_claim") == "yes":
        q_hint = "в посте есть р/с"
    elif p.get("has_account_claim") == "no":
        q_hint = "без счёта"
    t_hint = "есть блоки в тексте" if p.get("has_blocks_claim") else ""

    f_cell = ""
    if checklist.get("F_address") or checklist.get("F_director"):
        parts = []
        if checklist.get("F_address"):
            parts.append(checklist["F_address"])
        if checklist.get("F_director"):
            parts.append(str(checklist["F_director"]))
        f_cell = " | ".join(parts)

    inn = p.get("inn") or ""
    if not inn and p.get("inn_on_request"):
        inn = "по запросу"

    dossier = checklist.get("dossier") or ""
    if not dossier and companium and not companium.get("error"):
        # старые записи без dossier — коротко из счётчиков
        bits = []
        if companium.get("court_cases") is not None:
            n = companium["court_cases"]
            bits.append("суды: нет дел" if n == 0 else f"суды: {n}")
        if companium.get("enforcements") is not None:
            n = companium["enforcements"]
            bits.append("долги: нет" if n == 0 else f"долги: {n}")
        dossier = "; ".join(bits)

    employees = checklist.get("employees")
    if employees is None:
        employees = companium.get("employees")
    msp = checklist.get("msp") or companium.get("msp") or ""
    sanctions = checklist.get("sanctions") or companium.get("sanctions") or ""
    capital = checklist.get("capital_rub")
    if capital is None:
        capital = companium.get("capital_rub")
    taxes = checklist.get("taxes_rub")
    if taxes is None:
        taxes = companium.get("taxes_rub")
    founder = checklist.get("founder") or companium.get("founder") or ""

    from parser.extract import is_plausible_firm_name

    disp_name = ""
    if egrul.get("name"):
        disp_name = str(egrul["name"])
        if not disp_name.upper().startswith("ООО"):
            disp_name = f"ООО «{disp_name}»"
    elif p.get("name") and is_plausible_firm_name(str(p.get("name"))):
        disp_name = str(p.get("name"))
    elif inn and str(inn).isdigit():
        disp_name = f"ООО (ИНН {inn})"

    return [
        disp_name,
        inn or egrul.get("inn") or "",

        checklist.get("C_reg_date") or p.get("reg_date_raw") or "",
        _price_cell(p.get("price_rub")),
        "",  # цена покупки вручную
        f_cell,
        p.get("sno") or "",
        scoring.get("reg_before_2024") or "",
        _human_flag("I_reliable", checklist.get("I_reliable")),
        j_hint,
        checklist.get("K_loans_payables") or "",
        checklist.get("L_debts_il")
        or ("продавец: без долгов" if p.get("no_debts_claim") else ""),
        _human_flag("M_not_liquidating", checklist.get("M_not_liquidating")),
        _human_flag("N_not_excluding", checklist.get("N_not_excluding")),
        checklist.get("O_clean") or "",
        checklist.get("P_court_cases") or "",
        q_hint or "не сказано в посте",
        _r_cell(checklist, scoring, p),
        _zsk_cell_payload(p),
        t_hint or "не сказано в посте",
        _human_flag("U_reports_filed", checklist.get("U_reports_filed")) or "нет данных",
        _v_cell(checklist) or "нет данных",
        (p.get("listing_first_seen") or p.get("msg_date") or "")[:10],
        enrich.get("checked_at") or p.get("msg_date") or "",
        scoring.get("summary") or "нет оценки — запустите --rescore",
        p.get("link") or "",
        p.get("seller_username") or p.get("seller_from_msg") or "",
        scoring.get("score") or "",
        p.get("ogrn") or egrul.get("ogrn") or "",
        p.get("okved") or egrul.get("okved") or "",
        scoring.get("confidence") or "",
        checklist.get("status") or egrul.get("status") or "",
        "да" if p.get("is_duplicate") else "",
        dossier,
        employees if employees is not None else "",
        msp,
        sanctions,
        _money(capital),
        _money(taxes),
        founder,
        (p.get("raw_text") or "")[:2000],
    ]


_WIDTHS = {
    1: 28,
    2: 14,
    3: 14,
    4: 12,
    6: 36,
    7: 14,
    12: 22,
    15: 18,
    16: 14,
    19: 24,
    24: 48,
    25: 28,
    26: 16,
    33: 56,
    39: 28,
    40: 40,
}


def _first_seen_day(p: dict[str, Any]) -> str:
    """YYYY-MM-DD первого появления в чате (для листов)."""
    raw = p.get("listing_first_seen") or p.get("msg_date") or ""
    if not raw:
        return "без-даты"
    try:
        dt = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
        return dt.date().isoformat()
    except Exception:
        return str(raw)[:10] if len(str(raw)) >= 10 else "без-даты"


def _sheet_title(day: str) -> str:
    """Имя листа Excel <=31 символ: 07.08.2026."""
    if day == "без-даты":
        return "без даты"
    try:
        y, m, d = day.split("-")
        return f"{d}.{m}.{y}"
    except Exception:
        return day[:31]


def _write_sheet(ws: Worksheet, payloads: list[dict[str, Any]]) -> None:
    header_font = Font(bold=True)
    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, title)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # внутри дня: выше балл сверху
    rows = sorted(
        payloads,
        key=lambda p: int((p.get("scoring") or {}).get("score") or 0),
        reverse=True,
    )
    for r, p in enumerate(rows, 2):
        values = row_from_payload(p)
        for c, val in enumerate(values, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        zsk = (
            p.get("zsk_verified")
            or ((p.get("enrich") or {}).get("zsk_bot") or {}).get("level")
            or p.get("zsk_claim")
            or ""
        )
        # цвет всей строки по вердикту
        verdict = (p.get("scoring") or {}).get("verdict")
        if verdict in VERDICT_FILL:
            fill = VERDICT_FILL[verdict]
            for c in range(1, len(HEADERS) + 1):
                ws.cell(r, c).fill = fill
        # ЗСК поверх — чуть заметнее в своей колонке
        if zsk in ZSK_FILL:
            ws.cell(r, 19).fill = ZSK_FILL[zsk]

    for i in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = _WIDTHS.get(i, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(HEADERS))}{max(1, len(rows) + 1)}"
    )


def _has_real_inn(p: dict[str, Any]) -> bool:
    inn = (p.get("inn") or "").strip()
    if inn.isdigit() and len(inn) == 10:
        return True
    e_inn = str(((p.get("enrich") or {}).get("egrul") or {}).get("inn") or "").strip()
    return e_inn.isdigit() and len(e_inn) == 10


def export_xlsx(
    payloads: list[dict[str, Any]],
    path: Path,
    *,
    skip_duplicates: bool = True,
    by_first_seen_day: bool = True,
    require_inn: bool = True,
) -> Path:
    if skip_duplicates:
        payloads = [p for p in payloads if not p.get("is_duplicate")]
    if require_inn:
        payloads = [p for p in payloads if _has_real_inn(p)]

    wb = Workbook()
    # убрать дефолтный лист — пересоздадим
    default = wb.active

    if by_first_seen_day:
        groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for p in payloads:
            groups[_first_seen_day(p)].append(p)
        # слева свежие (новые даты), справа старые
        days = sorted(groups.keys(), reverse=True)
        first = True
        for day in days:
            title = _sheet_title(day)
            # уникальность имени листа
            base, n = title, 2
            while title in wb.sheetnames:
                title = f"{base}_{n}"[:31]
                n += 1
            if first:
                ws = default
                ws.title = title
                first = False
            else:
                ws = wb.create_sheet(title)
            _write_sheet(ws, groups[day])
        if not days:
            default.title = "пусто"
            _write_sheet(default, [])
    else:
        default.title = "Все лоты"
        _write_sheet(default, payloads)

    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(
            f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        )
        wb.save(alt)
        print(f"Excel занят ({path.name}) — сохранено как {alt.name}")
        return alt
