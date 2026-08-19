"""Выгрузка в CRM Lavok: локальный xlsx + JSON-пачки по HTTP/1.1 без прокси.

xlsx целиком с Windows рвёт TCP (10054). JSON по несколько строк — нет.
"""
from __future__ import annotations

import json
import os
import ssl
import subprocess
import tempfile
import time
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import ProxyHandler, Request, build_opener, HTTPSHandler

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import (
    CRM_EXPORT_PATH,
    DATA_DIR,
    LAVOK_INGEST_TOKEN,
    LAVOK_INGEST_URL,
)
from parser.export_apps_script import SHEET_HEADERS, build_export_body
from parser.export_fingerprint import fingerprint

# CRM ждёт «Отчетность» без ё
CRM_HEADERS = [h.replace("Отчётность", "Отчетность") for h in SHEET_HEADERS]

INN_COL = CRM_HEADERS.index("ИНН") + 1  # 1-based
SCORE_COL = CRM_HEADERS.index("Балл") + 1

HEADER_TO_FIELD = {
    "Источник": "source",
    "Название": "name",
    "ИНН": "inn",
    "Цена": "price",
    "Дата регистрации": "registered_at",
    "Налог": "tax",
    "Адрес и директор": "address_director",
    "Суды": "courts",
    "Долги / ИЛ": "debts",
    "Достоверность ЕГРЮЛ": "egrul_reliability",
    "Банкротство": "bankruptcy",
    "Обороты": "turnover",
    "Отчетность": "reporting",
    "Отчётность": "reporting",
    "Лизинг / залоги": "leasing",
    "ЗСК": "zsk",
    "Итог": "summary",
    "Балл": "score",
    "Первое появление": "first_seen",
    "Продавец": "seller",
    "Ссылка": "link",
    "Companium": "companium",
    "Статус ЕГРЮЛ": "egrul_status",
}

# curl на Windows стабильнее Python TLS к этому API
BATCH_SIZE = int(os.getenv("LAVOK_BATCH_SIZE", "15"))
SENT_PATH = DATA_DIR / "crm_sent.json"
FIELD_LIMITS = {
    "summary": 1200,
    "companium": 500,
    "address_director": 400,
    "zsk": 300,
    "courts": 200,
    "debts": 200,
    "turnover": 200,
    "leasing": 200,
    "name": 200,
}


def json_ingest_url(xlsx_url: str) -> str:
    url = (xlsx_url or "").strip().rstrip("/")
    if url.endswith("/ingest"):
        return f"{url}/json"
    if url.endswith("/ingest/json") or url.endswith("/ingest-json"):
        return url
    return f"{url}/ingest/json"


def _clip(rec: dict[str, Any]) -> dict[str, Any]:
    out = dict(rec)
    for key, lim in FIELD_LIMITS.items():
        val = out.get(key)
        if isinstance(val, str) and len(val) > lim:
            out[key] = val[: lim - 1] + "…"
    inn = str(out.get("inn") or "")
    out["inn"] = "".join(ch for ch in inn if ch.isdigit())
    if "score" in out:
        out["score"] = str(out["score"])
    return out


def items_from_body(body: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for spec in body.get("sheets") or []:
        sheet_date = str(spec.get("name") or "").strip()
        for row in spec.get("rows") or []:
            rec: dict[str, Any] = {"sheet_date": sheet_date}
            values = list(row or [])
            for idx, header in enumerate(CRM_HEADERS):
                field = HEADER_TO_FIELD.get(header)
                if not field or idx >= len(values):
                    continue
                val = values[idx]
                if val is None:
                    continue
                text = str(val).strip()
                if not text:
                    continue
                rec[field] = text
            if rec.get("inn"):
                items.append(_clip(rec))
    return items


def write_crm_xlsx(
    payloads: list[dict[str, Any]],
    path: Path | None = None,
    *,
    skip_duplicates: bool = True,
) -> tuple[Path, dict[str, Any], int]:
    """Пишет xlsx в формате CRM. Возвращает (path, body, skipped)."""
    path = path or CRM_EXPORT_PATH
    body, skipped = build_export_body(payloads, skip_duplicates=skip_duplicates)

    wb = Workbook()
    default = wb.active
    first = True
    for spec in body.get("sheets") or []:
        name = str(spec.get("name") or "лист")[:31]
        headers = list(CRM_HEADERS)
        rows = spec.get("rows") or []
        if first:
            ws = default
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)

        header_font = Font(bold=True)
        for c, title in enumerate(headers, 1):
            cell = ws.cell(1, c, title)
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                # ИНН / Балл — строго текст
                if c in (INN_COL, SCORE_COL) and val not in (None, ""):
                    s = str(val).replace("'", "")
                    if c == INN_COL:
                        s = "".join(ch for ch in s if ch.isdigit())
                    cell = ws.cell(r, c, s)
                    cell.number_format = "@"
                else:
                    cell = ws.cell(r, c, "" if val is None else val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 14
        ws.freeze_panes = "A2"

    if first:
        default.title = "пусто"
        for c, title in enumerate(CRM_HEADERS, 1):
            default.cell(1, c, title)

    path.parent.mkdir(parents=True, exist_ok=True)
    path = Path(path)
    try:
        wb.save(path)
    except PermissionError:
        from datetime import datetime

        alt = path.with_name(
            f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        )
        wb.save(alt)
        print(f"CRM xlsx занят ({path.name}) — сохранено как {alt.name}")
        path = alt

    n_rows = sum(len(s.get("rows") or []) for s in (body.get("sheets") or []))
    print(
        f"CRM xlsx: {path} | листов={len(body.get('sheets') or [])} | "
        f"строк={n_rows} | отброшено={skipped}"
    )
    return path, body, skipped


def _ssl_ctx() -> ssl.SSLContext:
    ctx = ssl.create_default_context()
    ctx.minimum_version = ssl.TLSVersion.TLSv1_2
    ctx.set_alpn_protocols(["http/1.1"])
    return ctx


def _opener():
    return build_opener(ProxyHandler({}), HTTPSHandler(context=_ssl_ctx()))


def _item_key(item: dict[str, Any]) -> str:
    return f"{item.get('inn')}|{item.get('sheet_date')}"


def _load_sent() -> dict[str, str]:
    if not SENT_PATH.is_file():
        return {}
    try:
        data = json.loads(SENT_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_sent(sent: dict[str, str]) -> None:
    SENT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SENT_PATH.write_text(
        json.dumps(sent, ensure_ascii=False, indent=0), encoding="utf-8"
    )


def _filter_delta(items: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    sent = _load_sent()
    out: list[dict[str, Any]] = []
    for it in items:
        h = fingerprint(it)
        if sent.get(_item_key(it)) == h:
            continue
        out.append(it)
    return out, len(items) - len(out)


def _mark_sent(items: list[dict[str, Any]]) -> None:
    sent = _load_sent()
    for it in items:
        sent[_item_key(it)] = fingerprint(it)
    _save_sent(sent)


def _post_json_curl(url: str, token: str, payload: dict[str, Any], timeout: int = 90) -> dict[str, Any]:
    raw = json.dumps(payload, ensure_ascii=False)
    with tempfile.NamedTemporaryFile(
        mode="w", encoding="utf-8", suffix=".json", delete=False
    ) as fh:
        fh.write(raw)
        tmp = fh.name
    try:
        cmd = [
            "curl.exe",
            "-sS",
            "-X",
            "POST",
            url,
            "-H",
            f"X-Lavok-Ingest-Token: {token}",
            "-H",
            "Content-Type: application/json; charset=utf-8",
            "-H",
            "Accept: application/json",
            "-H",
            "User-Agent: firmy-lavok-parser/1.3-curl",
            "--data-binary",
            f"@{tmp}",
            "--connect-timeout",
            "20",
            "--max-time",
            str(timeout),
            "--retry",
            "2",
            "--retry-delay",
            "1",
            "--noproxy",
            "*",
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
        out = (proc.stdout or "").strip()
        err = (proc.stderr or "").strip()
        if proc.returncode != 0:
            raise RuntimeError(f"curl exit={proc.returncode}: {err or out[:300]}")
        try:
            data = json.loads(out) if out else {}
        except Exception as exc:
            raise RuntimeError(f"curl bad json: {out[:300]}") from exc
        if isinstance(data, dict) and data.get("error"):
            raise RuntimeError(f"CRM error: {data}")
        return data if isinstance(data, dict) else {"ok": True, "data": data}
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass


def _post_json_urllib(url: str, token: str, payload: dict[str, Any], timeout: int = 60) -> dict[str, Any]:
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = Request(url, data=raw, method="POST")
    req.add_header("Content-Type", "application/json; charset=utf-8")
    req.add_header("X-Lavok-Ingest-Token", token)
    req.add_header("Accept", "application/json")
    req.add_header("Connection", "close")
    req.add_header("User-Agent", "firmy-lavok-parser/1.3")
    try:
        with _opener().open(req, timeout=timeout) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            status = getattr(resp, "status", 200)
    except HTTPError as exc:
        err_body = exc.read().decode("utf-8", errors="replace")
        try:
            data = json.loads(err_body) if err_body else {}
        except Exception:
            data = {"raw": err_body[:800]}
        raise RuntimeError(f"HTTP {exc.code}: {data}") from exc
    except URLError as exc:
        raise RuntimeError(str(exc.reason or exc)) from exc
    try:
        data = json.loads(body) if body else {}
    except Exception:
        data = {"raw": body[:800]}
    if status >= 400:
        raise RuntimeError(f"HTTP {status}: {data}")
    return data if isinstance(data, dict) else {"ok": True, "data": data}


def _post_json_requests(url: str, token: str, payload: dict[str, Any], timeout: int = 60) -> dict[str, Any]:
    import requests

    sess = requests.Session()
    sess.trust_env = False
    resp = sess.post(
        url,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "X-Lavok-Ingest-Token": token,
            "Accept": "application/json",
            "User-Agent": "firmy-lavok-parser/1.3",
            "Connection": "close",
        },
        json=payload,
        timeout=(20, timeout),
    )
    try:
        data = resp.json()
    except Exception:
        data = {"raw": (resp.text or "")[:800]}
    if resp.status_code >= 400:
        raise RuntimeError(f"HTTP {resp.status_code}: {data}")
    return data if isinstance(data, dict) else {"ok": True, "data": data}


def _post_json(url: str, token: str, payload: dict[str, Any], timeout: int = 90) -> dict[str, Any]:
    errors: list[str] = []
    # 1) curl — на Win чаще всего проходит туда, где Python ловит 10054
    try:
        return _post_json_curl(url, token, payload, timeout=timeout)
    except Exception as exc:
        errors.append(f"curl: {exc}")
    try:
        return _post_json_requests(url, token, payload, timeout=timeout)
    except Exception as exc:
        errors.append(f"requests: {exc}")
    try:
        return _post_json_urllib(url, token, payload, timeout=timeout)
    except Exception as exc:
        errors.append(f"urllib: {exc}")
    raise RuntimeError(" | ".join(errors))


def _post_batch(url: str, token: str, batch: list[dict[str, Any]], retries: int = 4) -> dict[str, Any]:
    last_err: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            return _post_json(url, token, {"items": batch})
        except Exception as exc:
            last_err = exc
            wait = min(20, 2 ** attempt)
            print(
                f"CRM ingest: fail {len(batch)} rows attempt {attempt}/{retries}: {exc} — sleep {wait}s",
                flush=True,
            )
            if attempt < retries:
                time.sleep(wait)
    raise RuntimeError(last_err)


def ingest_crm_body(
    body: dict[str, Any],
    *,
    token: str | None = None,
    url: str | None = None,
    force_all: bool = False,
) -> dict[str, Any]:
    xlsx_url = (url if url is not None else LAVOK_INGEST_URL).strip()
    token = (token if token is not None else LAVOK_INGEST_TOKEN).strip()
    if not xlsx_url:
        raise SystemExit("В .env нет LAVOK_INGEST_URL")
    if not token:
        raise SystemExit("В .env нет LAVOK_INGEST_TOKEN")

    endpoint = json_ingest_url(xlsx_url)
    items = items_from_body(body)
    if not items:
        print("CRM ingest: нечего слать (0 строк с ИНН)", flush=True)
        return {"sheets": 0, "upserted": 0, "created": 0, "updated": 0}

    skipped = 0
    if not force_all:
        items, skipped = _filter_delta(items)

    print(
        f"CRM ingest: POST {endpoint} | send={len(items)} | skipped_same={skipped} | "
        f"batch={BATCH_SIZE} | via=curl",
        flush=True,
    )
    if not items:
        print("CRM ingest: всё уже отправлялось — skip", flush=True)
        return {"sheets": 0, "upserted": 0, "created": 0, "updated": 0, "skipped": skipped}

    created = 0
    updated = 0
    upserted = 0
    sheets = 0
    failed: list[str] = []
    for offset in range(0, len(items), BATCH_SIZE):
        chunk = items[offset : offset + BATCH_SIZE]
        try:
            data = _post_batch(endpoint, token, chunk)
            _mark_sent(chunk)
        except Exception as exc:
            # пачка не прошла — по одной через curl
            if len(chunk) > 1:
                print(f"CRM ingest: дроблю {len(chunk)} → по 1", flush=True)
                for one in chunk:
                    try:
                        part = _post_batch(endpoint, token, [one])
                        _mark_sent([one])
                        created += int(part.get("created") or 0)
                        updated += int(part.get("updated") or 0)
                        upserted += int(part.get("upserted") or 0)
                        sheets = max(sheets, int(part.get("sheets") or 0))
                        print(
                            f"CRM ingest: batch ok inn={one.get('inn')}",
                            flush=True,
                        )
                    except Exception as one_exc:
                        print(
                            f"CRM ingest: SKIP inn={one.get('inn')}: {one_exc}",
                            flush=True,
                        )
                        failed.append(str(one.get("inn") or "?"))
                time.sleep(0.2)
                continue
            print(f"CRM ingest: SKIP inn={chunk[0].get('inn')}: {exc}", flush=True)
            failed.append(str(chunk[0].get("inn") or "?"))
            time.sleep(0.5)
            continue
        created += int(data.get("created") or 0)
        updated += int(data.get("updated") or 0)
        upserted += int(data.get("upserted") or 0)
        sheets = max(sheets, int(data.get("sheets") or 0))
        print(
            f"CRM ingest: batch {offset + 1}-{offset + len(chunk)}/{len(items)} ok",
            flush=True,
        )
        time.sleep(0.15)

    result = {
        "sheets": sheets,
        "upserted": upserted,
        "created": created,
        "updated": updated,
        "failed": len(failed),
        "skipped": skipped,
    }
    print(
        f"CRM ingest: ok | sheets={sheets} | upserted={upserted} | "
        f"created={created} | updated={updated} | failed={len(failed)} | "
        f"skipped_same={skipped}",
        flush=True,
    )
    if failed and upserted == 0 and skipped == 0:
        raise SystemExit(f"CRM ingest: все пачки упали ({len(failed)})")
    return result


def ingest_crm(
    path: Path,
    *,
    token: str | None = None,
    url: str | None = None,
    retries: int = 4,
) -> dict[str, Any]:
    """Совместимость: из xlsx больше не шлём."""
    raise SystemExit(
        "xlsx POST в CRM отключён (connection reset). "
        "Используй ingest_crm_body() / --export-crm"
    )


def export_and_ingest_crm(
    payloads: list[dict[str, Any]],
    *,
    skip_duplicates: bool = True,
    path: Path | None = None,
    upload: bool = True,
) -> Path:
    out, body, _skipped = write_crm_xlsx(
        payloads, path=path, skip_duplicates=skip_duplicates
    )
    if upload:
        ingest_crm_body(body)
    return out
